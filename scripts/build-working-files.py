#!/usr/bin/env python3
"""Build original B2B Playbook working-file workbooks. Run from repo root."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "templates"

NAVY = "1B3A4B"
INK = "1A1A1A"
MUTED = "5C6B73"
PAPER = "F7F4EF"
RULE = "D9D0C3"
INPUT = "FFF3C4"
TEACH = "E8F0E6"
WHITE = "FFFFFF"

font_title = Font(name="Calibri", size=16, bold=True, color=NAVY)
font_h = Font(name="Calibri", size=12, bold=True, color=NAVY)
font_label = Font(name="Calibri", size=11, bold=True, color=INK)
font_body = Font(name="Calibri", size=11, color=INK)
font_muted = Font(name="Calibri", size=10, italic=True, color=MUTED)
font_white = Font(name="Calibri", size=11, bold=True, color=WHITE)
fill_head = PatternFill("solid", fgColor=NAVY)
fill_paper = PatternFill("solid", fgColor=PAPER)
fill_input = PatternFill("solid", fgColor=INPUT)
fill_teach = PatternFill("solid", fgColor=TEACH)
fill_rule = PatternFill("solid", fgColor=RULE)
fill_white = PatternFill("solid", fgColor=WHITE)
wrap = Alignment(wrap_text=True, vertical="top")


def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def banner(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, text)
    c.font = font_title
    c.fill = fill_paper
    c.alignment = wrap
    ws.row_dimensions[row].height = 28


def note(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, text)
    c.font = font_muted
    c.alignment = wrap
    ws.row_dimensions[row].height = 36


def header_row(ws, row, labels):
    for i, lab in enumerate(labels, 1):
        c = ws.cell(row, i, lab)
        c.font = font_white
        c.fill = fill_head
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(labels))}{row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def input_cell(cell, value=None):
    if value is not None:
        cell.value = value
    cell.fill = fill_input
    cell.font = font_body
    cell.alignment = wrap


def build_demo():
    wb = Workbook()

    score = wb.active
    score.title = "Scorecard"
    banner(score, 1, 6, "B2B Playbook · Demo scorecard")
    note(
        score,
        2,
        6,
        "Score what happened, not how it felt. Discovery is the gate: if you cannot connect a click to a pain they said, do not show it. Yellow cells are inputs. Average formulas stay put. Copyright © 2026 Ivan Xu.",
    )
    score["A3"] = "Rep"
    score["B3"] = ""
    input_cell(score["B3"])
    score["C3"] = "Account"
    score["D3"] = ""
    input_cell(score["D3"])
    score["E3"] = "Date"
    score["F3"] = ""
    input_cell(score["F3"])
    score["A4"] = "Discovery complete?"
    score["B4"] = "yes / no / unknown"
    input_cell(score["B4"])
    score["C4"] = "Coach"
    score["D4"] = ""
    input_cell(score["D4"])

    header_row(score, 6, ["#", "Section", "Question", "Score 1–5 (blank = incomplete)", "Notes", "Section average"])

    rows = [
        (1, "People", "Were the required buying seats actually in the demo?"),
        (2, "People", "Did you reach people you had not met before the demo (not to pitch—to make the room less cold)?"),
        (3, "People", "Did you use names and pull each person into the conversation?"),
        (4, "Stage", "Did you recap their goals in their language?"),
        (5, "Stage", "Did you recap what is in the way?"),
        (6, "Stage", "Did you recap discovery out loud so they could correct you?"),
        (7, "Stage", "Did you paint a dated future they can see (not a feature list)?"),
        (8, "Stage", "Did you stake the meeting: sponsor the next room, or say we missed?"),
        (9, "Stage", "Was the stage block in customer language, not vendor language?"),
        (10, "Stage", "Did you offer one similar proof, then tee up how this helps them?"),
        (11, "During", "Did you start with the strongest relevant proof—not admin home?"),
        (12, "During", "Before each path: did you say why this click, tied to a pain they named?"),
        (13, "During", "Did you orient (what we are looking at) before each new surface?"),
        (14, "During", "So-what stack: did you reach a business consequence they own, not 'saves time'?"),
        (15, "During", "Did you only show what discovery earned?"),
        (16, "During", "After each key path: a short story they can remember?"),
        (17, "During", "After each key path: a real question (not 'any questions?')?"),
        (18, "End", "Did you ask what excited them most—and why?"),
        (19, "End", "If something fell flat, did you ask where you missed?"),
        (20, "End", "Did you leave with a dated next meeting (or a written no)?"),
    ]

    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv.error = "Use 1–5 or leave blank"
    dv.errorTitle = "Score"
    score.add_data_validation(dv)

    for i, (n, section, q) in enumerate(rows):
        r = 7 + i
        score.cell(r, 1, n).font = font_body
        score.cell(r, 2, section).font = font_label
        score.cell(r, 3, q).alignment = wrap
        score.cell(r, 3).font = font_body
        input_cell(score.cell(r, 4))
        dv.add(score.cell(r, 4))
        input_cell(score.cell(r, 5))
        score.row_dimensions[r].height = 32

    # Section averages: People 7-9, Stage 10-16, During 17-23, End 24-26
    score["F7"] = '=IF(COUNT(D7:D9)=3,ROUND(AVERAGE(D7:D9),2),"incomplete")'
    score.merge_cells("F7:F9")
    score["F10"] = '=IF(COUNT(D10:D16)=7,ROUND(AVERAGE(D10:D16),2),"incomplete")'
    score.merge_cells("F10:F16")
    score["F17"] = '=IF(COUNT(D17:D23)=7,ROUND(AVERAGE(D17:D23),2),"incomplete")'
    score.merge_cells("F17:F23")
    score["F24"] = '=IF(COUNT(D24:D26)=3,ROUND(AVERAGE(D24:D26),2),"incomplete")'
    score.merge_cells("F24:F26")
    for r in (7, 10, 17, 24):
        score.cell(r, 6).font = font_label
        score.cell(r, 6).alignment = Alignment(vertical="center")

    score["A28"] = "Total (all scored lines)"
    score["A28"].font = font_h
    score["D28"] = '=IF(COUNT(D7:D26)=20,ROUND(AVERAGE(D7:D26),2),"incomplete — score every line")'
    score["D28"].font = font_label
    score["A29"] = "Keep doing (top 3)"
    input_cell(score["B29"])
    score.merge_cells("B29:F29")
    score["A30"] = "Stop doing (top 3)"
    input_cell(score["B30"])
    score.merge_cells("B30:F30")
    score["A31"] = "One drill before the next demo"
    input_cell(score["B31"])
    score.merge_cells("B31:F31")
    for r in (29, 30, 31):
        score.row_dimensions[r].height = 40
        score.cell(r, 1).font = font_label

    score["A33"] = "Scale: 1 missed · 2 grazed · 3 weak attempt · 4 solid · 5 masterclass. An analyst stops at the number. A coach writes keep / stop / one drill."
    score["A33"].font = font_muted
    score.merge_cells("A33:F33")
    col_widths(score, [6, 12, 78, 22, 36, 22])

    q = wb.create_sheet("Questions")
    banner(q, 1, 2, "Do not ask “any questions?”")
    note(q, 2, 2, "Use these after a path and at the end. Rewrite in the customer’s words.")
    header_row(q, 4, ["When", "Ask"])
    questions = [
        ("After a path", "How does this compare to how you do this today?"),
        ("After a path", "You said X was the issue—where do you see this helping?"),
        ("After a path", "How would this change the weekly workflow?"),
        ("After a path", "Does this match how the team actually operates, or would something have to change?"),
        ("After a path", "How valuable would this be to the people who live in this job?"),
        ("Mid / end", "From what we covered, what stands out?"),
        ("Mid / end", "Who else on the team would feel this most?"),
        ("Mid / end", "How important is this capability to the objective you named?"),
        ("End", "What excited you most about what you saw?"),
        ("End", "Why did that resonate?"),
        ("End", "To what extent do you see this solving [the problem they named]?"),
        ("End", "How do you imagine the team using this in the first 30 days?"),
        ("If it landed", "It seemed to land—help me understand why, in your words."),
        ("If it missed", "That seemed flatter than I expected. Where did I miss?"),
    ]
    for i, (when, ask) in enumerate(questions):
        q.cell(5 + i, 1, when).font = font_label
        q.cell(5 + i, 2, ask).font = font_body
        q.cell(5 + i, 2).alignment = wrap
        q.row_dimensions[5 + i].height = 22
    col_widths(q, [16, 88])

    t = wb.create_sheet("Teaching fill")
    banner(t, 1, 3, "Invented example — not a customer. Delete this sheet before the file is your record.")
    t["A3"] = "Context"
    t["B3"] = "Mid-market analytics. Discovery: month-end close is a three-tool stitch. Champion + VP Finance. AE emailed the VP the day before."
    t.merge_cells("B3:C3")
    t["A4"] = "People"
    t["B4"] = "Both seats present. Silent director joined late and was never brought in — People average pulled down."
    t["A5"] = "Stage"
    t["B5"] = "Recap in their words. Dated future: March close in one sitting. Stake: sponsor CFO 30-min or say we missed."
    t["A6"] = "During"
    t["B6"] = "Opened on the close pack, not settings. So-what to “CFO stops asking for a side file.”"
    t["A7"] = "End"
    t["B7"] = "Excited: audit trail. Next meeting: security, Thursday 45 minutes."
    t["A8"] = "Drill"
    t["B8"] = "60-second name-and-role open so late joiners are not ghosts."
    for r in range(3, 9):
        t.cell(r, 1).font = font_label
        t.cell(r, 2).alignment = wrap
        t.cell(r, 2).fill = fill_teach
        t.row_dimensions[r].height = 36
        t.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    col_widths(t, [14, 70, 20])

    wb.save(OUT / "demo-scorecard.xlsx")


def build_lead_scoring():
    wb = Workbook()
    s = wb.active
    s.title = "Ledger"
    banner(s, 1, 7, "B2B Playbook · Lead-scoring ledger")
    note(
        s,
        2,
        7,
        "Write the action first, then the points. Fit and intent both belong. Do not score every open or visit. Changing a live model re-scores the database—log it. Yellow = inputs. Copyright © 2026 Ivan Xu.",
    )
    labels = [
        (3, "Threshold action (who / what / SLA)"),
        (4, "Hand-raiser rule (demo, pricing, contact)"),
        (5, "MQL definition"),
        (6, "SQL definition (human vs automatic)"),
        (7, "MAP / CRM of record"),
        (8, "Change-control note"),
    ]
    for r, lab in labels:
        s.cell(r, 1, lab).font = font_label
        s.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        input_cell(s.cell(r, 2))
        s.row_dimensions[r].height = 22
    header_row(s, 10, ["Attribute", "Fit or intent", "Sign", "Rule (interval / URL / form)", "Points (yours)", "Routes to", "Notes"])
    for r in range(11, 31):
        input_cell(s.cell(r, 1))
        input_cell(s.cell(r, 2))
        input_cell(s.cell(r, 3))
        input_cell(s.cell(r, 4))
        input_cell(s.cell(r, 5))
        input_cell(s.cell(r, 6))
        input_cell(s.cell(r, 7))
    dv_fit = DataValidation(type="list", formula1='"Fit,Intent,Negative,Exception"', allow_blank=True)
    dv_sign = DataValidation(type="list", formula1='"+,-"', allow_blank=True)
    s.add_data_validation(dv_fit)
    s.add_data_validation(dv_sign)
    dv_fit.add("B11:B30")
    dv_sign.add("C11:C30")
    col_widths(s, [28, 16, 10, 36, 16, 22, 36])

    f = wb.create_sheet("Forms by stage")
    banner(f, 1, 4, "Forms follow buying stage — not “more fields.”")
    note(f, 2, 4, "Progressive profiling: if you already have the field, ask the next one. Reuse forms; list off page + form.")
    header_row(f, 4, ["Stage", "Buyer job", "Fields we will ask", "Always-on (email, consent)"])
    stages = [
        ("Awareness", "I have a problem; I do not yet name the solution.", "", "Email + consent"),
        ("Consideration", "I can name the problem and I am comparing approaches.", "", "Email + consent"),
        ("Decision", "I am evaluating whether you can deliver.", "", "Email + consent"),
        ("Hand-raiser", "Demo / pricing / talk to sales — skip the leisurely ladder.", "", "Email + consent + phone if you will call"),
    ]
    for i, row in enumerate(stages):
        r = 5 + i
        f.cell(r, 1, row[0]).font = font_label
        f.cell(r, 2, row[1]).alignment = wrap
        input_cell(f.cell(r, 3))
        f.cell(r, 4, row[3]).font = font_body
        f.row_dimensions[r].height = 40
    col_widths(f, [16, 56, 48, 36])

    t = wb.create_sheet("Teaching fill")
    banner(t, 1, 2, "Invented example — not a customer. Delete before this is your MAP spec.")
    t["A3"] = "Action"
    t["B3"] = "Score ≥ 40 and fit = ICP → SDR task in 4 business hours. Demo form → same task immediately."
    t["A4"] = "Fit"
    t["B4"] = "Target industries, 50–2,000 employees, ops or IT role. Below 50: nurture, not SDR."
    t["A5"] = "Intent"
    t["B5"] = "Demo/pricing highest. Decision content next. Blog every 3 posts. No points per email open; clicks in intervals."
    t["A6"] = "Negative"
    t["B6"] = "Careers, student emails, employees. Customers blocked from hunter queue; CS owns them."
    t["A7"] = "SQL"
    t["B7"] = "Human accept in CRM. Not a second point gate."
    for r in range(3, 8):
        t.cell(r, 1).font = font_label
        t.cell(r, 2).fill = fill_teach
        t.cell(r, 2).alignment = wrap
        t.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
        t.row_dimensions[r].height = 36
    col_widths(t, [14, 100])
    wb.save(OUT / "lead-scoring-ledger.xlsx")


def build_demand():
    wb = Workbook()
    s = wb.active
    s.title = "Demand plan"
    banner(s, 1, 16, "B2B Playbook · Demand plan")
    note(
        s,
        2,
        16,
        "Each channel has its own conversion chain. Outputs are formulas. Do not type over them to make the year work. Yellow = inputs. Copyright © 2026 Ivan Xu.",
    )
    s["A3"] = "Output object (must match capacity + forecast)"
    input_cell(s["B3"], "SQO")
    s["C3"] = "Assumption freeze date"
    input_cell(s["D3"])
    s["E3"] = "Owner"
    input_cell(s["F3"])

    months = [f"M{i}" for i in range(1, 13)]
    header_row(s, 5, ["Channel", "Line", "Unit / rate"] + months)
    # Row structure per channel: volume, conv1, mid, conv2, output
    # Keep it simple: 5 channels × 5 lines

    channels = [
        ("Paid search", "Clicks", "Click → MQL", "MQL → SQL", "SQL → SQO"),
        ("Organic search", "Pageviews", "View → MQL", "MQL → SQL", "SQL → SQO"),
        ("Outbound SDR", "Accounts touched", "Touch → MQL", "MQL → SQL", "SQL → SQO"),
        ("Events", "Events held", "Leads / event", "Lead → MQL", "MQL → SQO"),
        ("Other (name it)", "Volume", "→ MQL", "MQL → SQL", "SQL → SQO"),
    ]

    r = 6
    channel_starts = []
    for ch, vol, c1, c2, c3 in channels:
        channel_starts.append(r)
        s.cell(r, 1, ch).font = font_label
        s.cell(r, 2, "Volume (input)").font = font_body
        s.cell(r, 3, vol).font = font_muted
        for c in range(4, 16):
            input_cell(s.cell(r, c), 0)
        r += 1
        s.cell(r, 1, ch)
        s.cell(r, 2, "Rate 1 (input)")
        s.cell(r, 3, c1)
        for c in range(4, 16):
            input_cell(s.cell(r, c), 0)
        r += 1
        s.cell(r, 1, ch)
        s.cell(r, 2, "Mid-count (formula)")
        s.cell(r, 3, "Volume × rate 1")
        for c in range(4, 16):
            vol_cell = f"{get_column_letter(c)}{r-2}"
            rate_cell = f"{get_column_letter(c)}{r-1}"
            s.cell(r, c, f"={vol_cell}*{rate_cell}")
            s.cell(r, c).number_format = "0.0"
        r += 1
        s.cell(r, 1, ch)
        s.cell(r, 2, "Rate 2 (input)")
        s.cell(r, 3, c2)
        for c in range(4, 16):
            input_cell(s.cell(r, c), 0)
        r += 1
        s.cell(r, 1, ch)
        s.cell(r, 2, "SQO (formula)")
        s.cell(r, 3, c3 + " — last rate × mid")
        for c in range(4, 16):
            mid = f"{get_column_letter(c)}{r-2}"
            rate = f"{get_column_letter(c)}{r-1}"
            s.cell(r, c, f"={mid}*{rate}")
            s.cell(r, c).number_format = "0.0"
            s.cell(r, c).font = font_label
        r += 1

    s.cell(r, 1, "All channels").font = font_h
    s.cell(r, 2, "SQO total").font = font_h
    sqo_rows = [start + 4 for start in channel_starts]
    for c in range(4, 16):
        parts = "+".join(f"{get_column_letter(c)}{rr}" for rr in sqo_rows)
        s.cell(r, c, f"={parts}")
        s.cell(r, c).font = font_h
        s.cell(r, c).number_format = "0.0"
        s.cell(r, c).fill = fill_rule
    r += 2
    s.cell(r, 1, "Paid cost this period (input, optional)").font = font_label
    for c in range(4, 16):
        input_cell(s.cell(r, c), 0)
    r += 1
    s.cell(r, 1, "Cost per SQO (formula)").font = font_label
    cost_row = r - 1
    total_row = r - 3
    for c in range(4, 16):
        cl = get_column_letter(c)
        s.cell(r, c, f'=IF({cl}{total_row}=0,"",{cl}{cost_row}/{cl}{total_row})')
        s.cell(r, c).number_format = "0"

    col_widths(s, [22, 22, 28] + [11] * 12)
    s.freeze_panes = "D6"

    t = wb.create_sheet("Teaching fill")
    banner(t, 1, 2, "Invented — not a forecast. Shows shape only. Your rates come from your last four quarters.")
    t["A3"] = "Story"
    t["B3"] = "Paid search and outbound are the only non-zero channels. Events: two shows. Do not copy these rates."
    t["A4"] = "Paid search example shape"
    t["B4"] = "Clicks 8,000; click→MQL 0.02; MQL→SQL 0.25; SQL→SQO 0.40. Cost in the cost row."
    t["A5"] = "Constraint reminder"
    t["B5"] = "If this sheet creates 15 SQO/month and capacity can work 12, tighten scoring—do not hire wishfully."
    for r in range(3, 6):
        t.cell(r, 1).font = font_label
        t.cell(r, 2).fill = fill_teach
        t.cell(r, 2).alignment = wrap
        t.row_dimensions[r].height = 40
    col_widths(t, [28, 90])
    wb.save(OUT / "gtm-demand-plan.xlsx")


def build_capacity():
    wb = Workbook()
    a = wb.active
    a.title = "Assumptions"
    banner(a, 1, 4, "B2B Playbook · Sales capacity — change inputs only")
    note(
        a,
        2,
        4,
        "Yellow cells drive every other sheet. Sample numbers on Teaching fill are invented. Stage names must match CRM and forecasting. Copyright © 2026 Ivan Xu.",
    )
    a["A4"] = "Model start (month 1)"
    input_cell(a["B4"])
    a["C4"] = "Owner"
    input_cell(a["D4"])
    a["A5"] = "Output object (must match demand plan)"
    input_cell(a["B5"], "Closed won")

    header_row(a, 7, ["Input", "Segment A (e.g. MM)", "Segment B (e.g. Ent)", "Notes"])
    inputs = [
        "Annual quota / AE",
        "Expected attainment (0–1)",
        "Ramp months",
        "Average first-contract value",
        "Stage 1 name",
        "Stage 2 name",
        "Stage 3 name",
        "Stage 4 (won) name",
        "Conv stage1 → 2",
        "Conv stage2 → 3",
        "Conv stage3 → won",
        "Stage 1 length (months)",
        "Stage 2 length (months)",
        "Stage 3 length (months)",
        "AEs supported per SDR",
        "AEs supported per SE",
        "AEs supported per manager",
        "Contract term (months)",
        "Logo churn at renewal (0–1)",
        "Upsell on retained (0–1)",
        "Customers at model start",
    ]
    for i, lab in enumerate(inputs):
        r = 8 + i
        a.cell(r, 1, lab).font = font_label
        input_cell(a.cell(r, 2))
        input_cell(a.cell(r, 3))
        input_cell(a.cell(r, 4))
    a["A30"] = "NRR (formula, segment A)"
    a["B30"] = "=(1-B26)*(1+B27)"
    a["C30"] = "=(1-C26)*(1+C27)"
    a["A30"].font = font_label
    a["B30"].number_format = "0.00"
    a["C30"].number_format = "0.00"
    a["A31"] = "Note: NRR here is a simple (1−churn)×(1+upsell) sketch, not a cohort engine."
    a["A31"].font = font_muted
    a.merge_cells("A31:D31")
    col_widths(a, [36, 24, 24, 40])

    cap = wb.create_sheet("AE roster")
    banner(cap, 1, 8, "List people. Start month and ramp are inputs. Payroll ≠ quota-bearing.")
    header_row(
        cap,
        3,
        [
            "Segment A or B",
            "Name",
            "Annual quota",
            "Start month (1–12)",
            "Ramp months",
            "M1 on payroll (1/0)",
            "M1 quota-bearing (1/0)",
            "Notes",
        ],
    )
    for r in range(4, 16):
        for c in range(1, 9):
            input_cell(cap.cell(r, c))
    cap["A17"] = "Payroll count"
    cap["B17"] = "=COUNTA(B4:B15)"
    cap["A18"] = "Quota-bearing this snapshot (sum of M1 quota-bearing)"
    cap["B18"] = "=SUM(G4:G15)"
    cap["A17"].font = font_label
    cap["A18"].font = font_label
    col_widths(cap, [16, 22, 16, 18, 14, 18, 22, 36])

    w = wb.create_sheet("Waterfall")
    banner(w, 1, 14, "Stage-to-stage. Put duration in your head: wins do not land in the month you created stage 1.")
    note(w, 2, 14, "Yellow = volume in at stage 1. Conversions reference Assumptions. This is a 12-month sketch, not a delay engine—shift wins by stage length in your own calendar if cycle > 1 month.")
    header_row(w, 4, ["Line"] + [f"M{i}" for i in range(1, 13)])
    w["A5"] = "Stage 1 in (input)"
    for c in range(2, 14):
        input_cell(w.cell(5, c), 0)
    w["A6"] = "→ Stage 2 (formula, uses Assumptions conv A)"
    for c in range(2, 14):
        cl = get_column_letter(c)
        w.cell(6, c, f"={cl}5*Assumptions!B16")
        w.cell(6, c).number_format = "0.0"
    w["A7"] = "→ Stage 3"
    for c in range(2, 14):
        cl = get_column_letter(c)
        w.cell(7, c, f"={cl}6*Assumptions!B17")
        w.cell(7, c).number_format = "0.0"
    w["A8"] = "→ Won (same month sketch)"
    for c in range(2, 14):
        cl = get_column_letter(c)
        w.cell(8, c, f"={cl}7*Assumptions!B18")
        w.cell(8, c).number_format = "0.0"
        w.cell(8, c).font = font_label
    w["A10"] = "Won × segment A ACV (formula)"
    for c in range(2, 14):
        cl = get_column_letter(c)
        w.cell(10, c, f"={cl}8*Assumptions!B11")
        w.cell(10, c).number_format = "$#,##0"
    w["A12"] = "Reconcile in one sentence (input): demand SQOs vs these wins vs roster capacity."
    w.merge_cells("A12:N12")
    input_cell(w["A13"])
    w.merge_cells("A13:N13")
    w.row_dimensions[13].height = 36
    col_widths(w, [42] + [11] * 12)

    rev = wb.create_sheet("Renewals")
    banner(rev, 1, 5, "Retention is a parent of the number. New logo cannot silently cover a leak.")
    header_row(rev, 3, ["Item", "Segment A", "Segment B", "Formula / note", ""])
    rev["A4"] = "Starting customers"
    rev["B4"] = "=Assumptions!B28"
    rev["C4"] = "=Assumptions!C28"
    rev["A5"] = "Up for renewal this year (input if not all)"
    input_cell(rev["B5"])
    input_cell(rev["C5"])
    rev["A6"] = "Expected retained logos"
    rev["B6"] = "=IF(B5=\"\",\"\",B5*(1-Assumptions!B26))"
    rev["C6"] = "=IF(C5=\"\",\"\",C5*(1-Assumptions!C26))"
    rev["A7"] = "NRR sketch"
    rev["B7"] = "=Assumptions!B30"
    rev["C7"] = "=Assumptions!C30"
    rev["B7"].number_format = "0.00"
    rev["C7"].number_format = "0.00"
    for r in range(4, 8):
        rev.cell(r, 1).font = font_label
    rev["A9"] = "If NRR < 1, write how much new logo must cover the leak:"
    rev.merge_cells("A9:C9")
    input_cell(rev["A10"])
    rev.merge_cells("A10:C10")
    rev.row_dimensions[10].height = 36
    col_widths(rev, [36, 16, 16, 40, 12])

    t = wb.create_sheet("Teaching fill")
    banner(t, 1, 2, "Invented — not your quota. Clear Assumptions before this file is a board artifact.")
    t["A3"] = "Story"
    t["B3"] = "4 MM AEs (one ramping), 2 enterprise (one starts month 7). Attainment from last year 0.75 / 0.70. 40 customers already on the book."
    t["A4"] = "Constraint"
    t["B4"] = "Demand creates ~15 SQO/month; capacity works ~12. Do not hire AE five."
    t["A5"] = "Do not copy"
    t["B5"] = "Any $ quota, 78% attainment, or 5-year logo table you saw in a sample model."
    for r in range(3, 6):
        t.cell(r, 1).font = font_label
        t.cell(r, 2).fill = fill_teach
        t.cell(r, 2).alignment = wrap
        t.row_dimensions[r].height = 40
    col_widths(t, [14, 100])
    wb.save(OUT / "gtm-sales-capacity.xlsx")


def _object_sheet(wb, title):
    ws = wb.create_sheet(title)
    banner(ws, 1, 9, f"B2B Playbook · {title} field map")
    note(ws, 2, 9, "Left = legacy. Right = new org. Types must match. Yellow = inputs. Copyright © 2026 Ivan Xu.")
    header_row(
        ws,
        4,
        [
            "Legacy label",
            "Legacy API name",
            "Legacy type",
            "Populated records",
            "Keep / kill / transform",
            "New label",
            "New API name",
            "New type",
            "Standard or custom",
        ],
    )
    for r in range(5, 35):
        for c in range(1, 10):
            input_cell(ws.cell(r, c))
    col_widths(ws, [22, 22, 16, 18, 18, 22, 22, 16, 18])
    return ws


def build_crm_map():
    wb = Workbook()
    s = wb.active
    s.title = "Scope"
    banner(s, 1, 4, "B2B Playbook · CRM field map")
    note(
        s,
        2,
        4,
        "Map left to right. Count populated records before you keep a field. Duplicate an object tab for Task, Contract, or custom objects. Copyright © 2026 Ivan Xu.",
    )
    labels = [
        (4, "From-org"),
        (5, "To-org"),
        (6, "Freeze date"),
        (7, "Owner"),
        (8, "Objects we will not move"),
        (9, "Population rule (below n → kill)"),
    ]
    for r, lab in labels:
        s.cell(r, 1, lab).font = font_label
        s.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        input_cell(s.cell(r, 2))
    col_widths(s, [36, 28, 20, 20])

    for name in ("Lead", "Account", "Contact", "Opportunity"):
        _object_sheet(wb, name)

    p = wb.create_sheet("Picklists")
    banner(p, 1, 4, "Every picklist and multi-select. Stage values must match the forecast page.")
    header_row(p, 3, ["Object", "Field", "Value", "Notes"])
    for r in range(4, 40):
        for c in range(1, 5):
            input_cell(p.cell(r, c))
    col_widths(p, [18, 28, 28, 40])

    t = wb.create_sheet("Teaching fill")
    banner(t, 1, 2, "Invented — not your schema. Delete before this is the migration spec.")
    t["A3"] = "Keep"
    t["B3"] = "Lead email, company, source, score, owner. Opportunity amount, close date, stage, next step, forecast category."
    t["A4"] = "Kill"
    t["B4"] = "Unused UTM customs with almost no populated rows. Vibe field named Temperature."
    t["A5"] = "Picklist"
    t["B5"] = "Stages = the five names on the forecast page. Retire Open / Working / Hot."
    for r in range(3, 6):
        t.cell(r, 1).font = font_label
        t.cell(r, 2).fill = fill_teach
        t.cell(r, 2).alignment = wrap
        t.row_dimensions[r].height = 40
    col_widths(t, [12, 100])
    wb.save(OUT / "crm-field-map.xlsx")


def build_vendor_eval():
    wb = Workbook()
    s = wb.active
    s.title = "Jobs"
    banner(s, 1, 6, "B2B Playbook · Vendor evaluation")
    note(
        s,
        2,
        6,
        "Split jobs before vendors. Community scores are prompts, not your score. Yellow = inputs. Copyright © 2026 Ivan Xu.",
    )
    s["A4"] = "Bake-off owner"
    input_cell(s["B4"])
    s["C4"] = "Review / kill date"
    input_cell(s["D4"])
    s["A5"] = "Required CRM writes (two-way?)"
    s.merge_cells("B5:F5")
    input_cell(s["B5"])
    header_row(
        s,
        7,
        [
            "Job (cadence / CI / forecast / other)",
            "Who lives in it weekly",
            "Must write to CRM",
            "In this buy? (yes/no)",
            "Why / why not",
            "Notes",
        ],
    )
    for r in range(8, 14):
        for c in range(1, 7):
            input_cell(s.cell(r, c))
    col_widths(s, [36, 22, 22, 16, 40, 28])

    v = wb.create_sheet("Bake-off")
    banner(v, 1, 8, "One row per vendor you will actually talk to. Do not paste a published 9.2/10.")
    header_row(
        v,
        3,
        [
            "Vendor",
            "Jobs it claims",
            "Win reasons (their buyers / our tests)",
            "Opportunity areas we will test",
            "Pricing model",
            "Time-to-live-in-tool (our range)",
            "Two-way CRM sync (yes/no/unknown)",
            "Alternatives they were compared to",
        ],
    )
    for r in range(4, 10):
        for c in range(1, 9):
            input_cell(v.cell(r, c))
            v.row_dimensions[r].height = 36
    col_widths(v, [18, 22, 36, 36, 20, 24, 22, 28])

    t = wb.create_sheet("Teaching fill")
    banner(t, 1, 2, "Invented — not a ranking. Delete before procurement.")
    t["A3"] = "Jobs"
    t["B3"] = "Cadence this year. Conversation intelligence later. Forecast stays in CRM."
    t["A4"] = "Tests"
    t["B4"] = "Two-way opportunity sync, cadence builder in our objects, support in our hours. Not a borrowed satisfaction score."
    t["A5"] = "Kill"
    t["B5"] = "If required fields cannot sync two-way, it is a no—not phase two."
    for r in range(3, 6):
        t.cell(r, 1).font = font_label
        t.cell(r, 2).fill = fill_teach
        t.cell(r, 2).alignment = wrap
        t.row_dimensions[r].height = 40
    col_widths(t, [12, 100])
    wb.save(OUT / "vendor-evaluation.xlsx")


if __name__ == "__main__":
    OUT.mkdir(exist_ok=True)
    build_demo()
    build_lead_scoring()
    build_demand()
    build_capacity()
    build_crm_map()
    build_vendor_eval()
    print("wrote", sorted(p.name for p in OUT.glob("*.xlsx")))
